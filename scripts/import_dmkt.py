"""Parse documents/*.xlsx ĐMKT → shared catalog + seed lệnh SX (GitHub demo)."""
from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "documents"
OUT_CATALOG = ROOT / "shared" / "src" / "data" / "dmktCatalog.ts"
OUT_ORDERS = ROOT / "backend" / "src" / "data" / "seedOrders.ts"

TEAM_NAME = {
    "t_hot": "Tổ Dập nóng – P.V.Chí",
    "t_auto": "Tổ Tự động – P.V.Sang",
    "t_asm": "Tổ Lắp ráp – N.T.Hoa",
}
STAGE = {"t_hot": "hot_forge", "t_auto": "auto", "t_asm": "assembly"}
MG = {"t_hot": "mg_hot", "t_auto": "mg_auto", "t_asm": "mg_asm"}

PRODUCT_RULES = [
    ("van cua", dict(id="p2", code="NOVO-VC-20", name="Van cửa NOVO 20")),
    ("kuma", dict(id="p4", code="KUMA-15-TKM", name="Vòi KUMA 15 tay gạt K1 xanh TKM")),
    ("voi", dict(id="p4", code="KUMA-15-TKM", name="Vòi KUMA 15 tay gạt K1 xanh TKM")),
    ("van bi", dict(id="p3", code="NOVO-VB-15-ABS", name="Van bi NOVO 15 tay ABS")),
    ("tay hk", dict(id="p6", code="NOVO-VG-15-LH", name="Van góc LH 1C NOVO 15 tay hợp kim")),
    ("hop kim", dict(id="p6", code="NOVO-VG-15-LH", name="Van góc LH 1C NOVO 15 tay hợp kim")),
    ("khoa", dict(id="p5", code="NOVO-VG-15-KHOA", name="Van góc 1C sau ĐH NOVO 15 tay khóa")),
    ("abs", dict(id="p1", code="NOVO-VG-15", name="Van góc 1C sau ĐH NOVO 15 tay ABS")),
]


def fold(s: str) -> str:
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().lower()


def slug(s: str, n: int = 20) -> str:
    s = fold(s)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return (s or "x")[:n].upper() or "X"


PART_CODE_KEYS = [
    ("lap rap", "ASM"),
    ("bao goi", "PACK"),
    ("dai oc 2", "DAIOC2"),
    ("dai oc", "DAIOC"),
    ("oc ap", "OCAL"),
    ("oc dem", "OCDEM"),
    ("dem ham", "DEMHAM"),
    ("dem truc", "DEMTRUC"),
    ("vong dem", "VONGDEM"),
    ("dau voi", "DAU"),
    ("nap", "NAP"),
    ("than", "THAN"),
    ("truc", "TRUC"),
    ("dia", "DIA"),
]


def part_code_of(pcode: str, part_name: str, idx: int, used: set[str]) -> str:
    n = fold(part_name)
    suffix = f"{idx:02d}"
    for key, tag in PART_CODE_KEYS:
        if key in n:
            suffix = tag
            break
    code = f"{pcode}-{suffix}"
    if code in used:
        code = f"{pcode}-{suffix}-{idx:02d}"
    used.add(code)
    return code


def cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(v)
    return str(v).replace("\n", " ").strip()


def iter_sheets(path: Path):
    if path.suffix.lower() == ".xls":
        import xlrd

        book = xlrd.open_workbook(path)
        for name in book.sheet_names():
            sh = book.sheet_by_name(name)
            rows = [
                [cell(sh.cell_value(r, c)) for c in range(sh.ncols)]
                for r in range(sh.nrows)
            ]
            yield name, rows
        return
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([cell(c) for c in row])
        yield name, rows
    wb.close()


def is_mau_van(name: str) -> bool:
    n = fold(name)
    if "cham cong" in n or "ao phong" in n:
        return False
    if "tieu hao" in n or "chi tiet" in n:
        return False
    if "dong goi" in n or "bao goi" in n:
        return False
    return "mau van" in n or n.strip() == "mau van"


def is_pack(name: str) -> bool:
    n = fold(name)
    return "dong goi" in n or "bao goi" in n or "mau voi" in n


def skip_part_row(part: str, process: str) -> bool:
    blob = fold(f"{part} {process}")
    if not process:
        return True
    keys = (
        "tong don gia",
        "cong don gia",
        "so cong cho",
        "don gia gia cong",
        "ghi chu",
        "ngay",
        "tong giam doc",
        "ap dung",
    )
    return any(k in blob for k in keys)


def team_of(code: str, process: str = "") -> str:
    raw = fold(f"{code} {process}")
    if any(k in raw for k in ("lap rap", "bao goi", "pk", "phu kien")):
        return "t_asm"
    if any(k in raw for k in ("td", "cnc", "tu dong", "danh bong", "hoan thien")):
        return "t_auto"
    if any(k in raw for k in ("cp", "dap", "cat phoi", "hot")):
        return "t_hot"
    if "to lap" in raw:
        return "t_asm"
    return "t_hot"


def find_header(rows: list[list[str]]) -> int:
    for i, row in enumerate(rows):
        if fold(row[0] if row else "") == "tt":
            return i
    return -1


def title_from(rows: list[list[str]]) -> str:
    for row in rows[:8]:
        t = " ".join(row).strip()
        ft = fold(t)
        if "qtcn" in ft or "tien luong" in ft:
            t = re.sub(
                r"^.*?GIA CÔNG\s+",
                "",
                t,
                flags=re.I,
            )
            t = re.sub(r"^.*?LẮP RÁP VÀ BAO GÓI\s+", "", t, flags=re.I)
            return t.strip(" -")
    return ""


def parse_process_sheet(rows: list[list[str]], assembly: bool) -> list[dict]:
    hi = find_header(rows)
    if hi < 0:
        return []
    out: list[dict] = []
    part = ""
    seq = 0
    for row in rows[hi + 1 :]:
        while len(row) < 12:
            row.append("")
        tt, part_cell, process, team, people, machine, _jig, tech, _uom, quota = row[:10]
        if skip_part_row(part_cell, process) and not (assembly and process):
            if skip_part_row(part_cell, process):
                continue
        if part_cell and not fold(part_cell).startswith("tong") and "don gia" not in fold(part_cell):
            part = re.sub(r"^\d+[:.)]\s*", "", part_cell).strip()
            seq = 0
        if not process or not part:
            continue
        if skip_part_row(part, process):
            continue
        seq += 1
        tid = team_of(team, process)
        try:
            quota_n = float(str(quota).replace(",", ".")) if str(quota).strip() else 0
        except ValueError:
            quota_n = 0
        try:
            people_n = float(str(people).replace(",", ".")) if str(people).strip() else 0
        except ValueError:
            people_n = 0
        out.append(
            {
                "partName": part,
                "process": process.strip(),
                "processSeq": seq,
                "teamId": tid,
                "machine": machine.strip(),
                "techNote": tech.strip(),
                "quota": quota_n,
                "people": people_n,
            }
        )
    return out


def match_product(filename: str) -> dict:
    n = fold(filename)
    for key, meta in PRODUCT_RULES:
        if key in n:
            return dict(meta)
    raise SystemExit(f"Không map được file: {filename}")


def specs_from_note(note: str) -> dict:
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*mm", fold(note))
    if m:
        return {"A": "float", "B": "boolean"}
    if note:
        return {"A": "text", "B": "boolean"}
    return {"A": "float", "B": "boolean"}


def spec_cols(note: str) -> list[str]:
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*mm", note, re.I)
    if m:
        return [m.group(1).replace(",", "."), "NQ"] + [""] * 9
    if "bản vẽ" in fold(note) or "ban ve" in fold(note):
        return ["Theo BV", "NQ"] + [""] * 9
    return ["NQ"] + [""] * 10


def ts_json(value) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2)


def write_ts(path: Path, body: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(body.replace("\r\n", "\n"), encoding="utf-8")
    print("wrote", path.relative_to(ROOT), "bytes", path.stat().st_size)


def main() -> None:
    files = sorted(p for p in DOCS.iterdir() if p.suffix.lower() in {".xlsx", ".xls"})
    if not files:
        raise SystemExit(f"Không thấy file ĐMKT trong {DOCS}")

    products: list[dict] = []
    semis: list[dict] = []
    boms: list[dict] = []
    processes: list[dict] = []
    machines_map: dict[str, dict] = {}
    warehouse: list[dict] = []
    seen_product: set[str] = set()
    catalog_by_product: dict[str, list[dict]] = {}

    for path in files:
        meta = match_product(path.name)
        pid, pcode, pname = meta["id"], meta["code"], meta["name"]
        print("file", path.name, "->", pcode)

        mau_rows: list[dict] = []
        pack_rows: list[dict] = []
        title = pname
        for sheet, rows in iter_sheets(path):
            if is_mau_van(sheet):
                mau_rows.extend(parse_process_sheet(rows, False))
            elif is_pack(sheet):
                pack_rows.extend(parse_process_sheet(rows, True))

        if pid not in seen_product:
            seen_product.add(pid)
            products.append(
                {
                    "id": pid,
                    "code": pcode,
                    "name": pname,
                    "description": f"Import ĐMKT {path.name}",
                    "unitOfMeasureId": "uom-pcs",
                    "active": True,
                }
            )
            warehouse.append(
                {
                    "id": f"ws-{pid}",
                    "warehouseId": "wh-main",
                    "itemKind": "product",
                    "itemId": pid,
                    "qty": 15,
                }
            )

        parts: dict[str, list[dict]] = {}
        for row in mau_rows + pack_rows:
            parts.setdefault(row["partName"], []).append(row)

        product_semis: list[dict] = []
        used_codes: set[str] = set()
        for idx, (part_name, steps) in enumerate(parts.items(), start=1):
            sid = f"sp-{pcode.lower()}-{idx:02d}"
            scode = part_code_of(pcode, part_name, idx, used_codes)
            note0 = next((s["techNote"] for s in steps if s["techNote"]), "")
            semi = {
                "id": sid,
                "code": scode,
                "name": part_name,
                "productId": pid,
                "unitOfMeasureId": "uom-pcs",
                "measurementSpecs": specs_from_note(note0),
                "active": True,
            }
            semis.append(semi)
            product_semis.append({"semi": semi, "steps": steps})
            bom_id = f"bom-{sid}"
            boms.append({"id": bom_id, "name": f"BOM {part_name}", "semiProductId": sid})
            warehouse.append(
                {
                    "id": f"ws-{sid}",
                    "warehouseId": "wh-main",
                    "itemKind": "semi_product",
                    "itemId": sid,
                    "qty": 40 + (idx * 7) % 80,
                }
            )
            for step in steps:
                tid = step["teamId"]
                processes.append(
                    {
                        "id": f"bp-{sid}-{step['processSeq']}",
                        "bomId": bom_id,
                        "name": step["process"],
                        "productionTeamId": tid,
                        "machineGroupId": MG[tid],
                        "quotaPerShift": int(step["quota"] or 0),
                        "sortOrder": step["processSeq"],
                    }
                )
                mname = step["machine"]
                if mname:
                    mid = "m-" + slug(mname, 16).lower()
                    if mid not in machines_map:
                        machines_map[mid] = {
                            "id": mid,
                            "name": mname,
                            "accountingCode": slug(mname, 12),
                            "code": slug(mname, 12),
                            "machineGroupId": MG[tid],
                            "productionTeamId": tid,
                            "teamId": tid,
                            "specs": {},
                            "active": True,
                        }
        catalog_by_product[pid] = product_semis

    products.sort(key=lambda p: p["id"])
    machines = list(machines_map.values())
    machines.sort(key=lambda m: m["name"])

    catalog_body = f"""/** Auto-generated from documents/*.xlsx — `python scripts/import_dmkt.py` */
import type {{
  Bom,
  BomProcess,
  Machine,
  Product,
  SemiProduct,
  WarehouseStock,
}} from "../types/index.js";

export const DMKT_PRODUCTS: Product[] = {ts_json(products)};

export const DMKT_SEMI: SemiProduct[] = {ts_json(semis)};

export const DMKT_BOMS: Bom[] = {ts_json(boms)};

export const DMKT_BOM_PROCESSES: BomProcess[] = {ts_json(processes)};

export const DMKT_MACHINES: Machine[] = {ts_json(machines)};

export const DMKT_WAREHOUSE: WarehouseStock[] = {ts_json(warehouse)};
"""
    write_ts(OUT_CATALOG, catalog_body)

    def jobs_for(
        pid: str,
        qty: int,
        *,
        limit_parts: int | None = None,
        assign_first: bool = False,
        progress: bool = False,
        keep_b1: bool = False,
    ) -> list[dict]:
        chunks = catalog_by_product[pid]
        if limit_parts is not None:
            chunks = chunks[:limit_parts]
        jobs: list[dict] = []
        n = 0
        for chunk in chunks:
            semi = chunk["semi"]
            for step in chunk["steps"]:
                n += 1
                tid = step["teamId"]
                assigned = assign_first and n <= 2
                workers = ["Cường 2T3"] if assigned else []
                assignments = (
                    [{"workerId": "u6", "workerName": "Cường 2T3", "machineName": step["machine"] or "CP"}]
                    if assigned
                    else []
                )
                done = 0
                fail = 0
                status = "unassigned"
                if assigned:
                    status = "in_progress"
                    if progress:
                        done = int(qty * 0.6)
                        fail = 4
                elif not assign_first:
                    status = "unassigned"
                job = {
                    "id": "b1" if keep_b1 and n == 1 else f"b-{pid}-{semi['id']}-{step['processSeq']}",
                    "bomCode": f"BOM-{semi['code']}-{step['processSeq']:02d}",
                    "partCode": semi["code"],
                    "partName": semi["name"],
                    "partGroup": semi["name"],
                    "processSeq": step["processSeq"],
                    "rawMaterial": step["techNote"] or "",
                    "machine": step["machine"] or "",
                    "process": step["process"],
                    "targetQty": qty,
                    "passQty": done,
                    "failQty": fail,
                    "assignedTeamId": tid,
                    "assignedTeamName": TEAM_NAME[tid],
                    "assignedWorkers": workers,
                    "workerAssignments": assignments,
                    "processStage": STAGE[tid],
                    "status": status,
                    "quota": str(int(step["quota"])) if step["quota"] else "",
                    "specCols": spec_cols(step["techNote"]),
                    "techNote": step["techNote"],
                    "semiProductId": semi["id"],
                    "attachments": [],
                    "workerEntries": [],
                }
                if keep_b1 and n == 1:
                    job["specCols"] = ["Ø20", "Ø9", "1.5", "M6", "Ø4", "NQ"] + [""] * 5
                    job["attachments"] = [
                        {
                            "id": "ba1",
                            "name": "ThongSoKyThuat_NOVO20.xlsx",
                            "type": "excel",
                            "size": "540 KB",
                            "uploadedBy": "Nguyễn Văn An",
                            "uploadedAt": "15/01/2024",
                        }
                    ]
                    job["techNote"] = step["techNote"] or "Lắp ghép ren theo tiêu chuẩn. SP không bavia."
                jobs.append(job)
        return jobs

    p1 = next(p for p in products if p["id"] == "p1")
    p2 = next(p for p in products if p["id"] == "p2")
    p3 = next((p for p in products if p["id"] == "p3"), None)
    p4 = next((p for p in products if p["id"] == "p4"), None)
    p5 = next((p for p in products if p["id"] == "p5"), None)

    orders = [
        {
            "id": "o1",
            "orderNo": "LSX-2024-001",
            "productId": "p1",
            "productCode": p1["code"],
            "productLine": p1["name"],
            "customer": "Nội bộ",
            "targetQty": 500,
            "createdBy": "Nguyễn Văn An",
            "createdAt": "15/01/2024",
            "deadline": "2024-03-01",
            "priority": "high",
            "status": "in_progress",
            "pendingApproval": False,
            "note": "Lệnh demo đo điểm — dữ liệu ĐMKT van góc ABS",
            "attachments": [
                {
                    "id": "a2",
                    "name": "ThongSoKyThuat_NOVO20.xlsx",
                    "type": "excel",
                    "size": "540 KB",
                    "uploadedBy": "Nguyễn Văn An",
                    "uploadedAt": "15/01/2024",
                }
            ],
            "boms": jobs_for("p1", 500, limit_parts=2, assign_first=True, progress=True, keep_b1=True),
        },
        {
            "id": "o2",
            "orderNo": "LSX-2024-002",
            "productId": "p2",
            "productCode": p2["code"],
            "productLine": p2["name"],
            "customer": "Khách hàng A",
            "targetQty": 350,
            "createdBy": "Trần Thị Bình",
            "createdAt": "16/01/2024",
            "deadline": "2024-02-05",
            "priority": "normal",
            "status": "approved",
            "pendingApproval": False,
            "attachments": [],
            "boms": jobs_for("p2", 350, limit_parts=4),
        },
        {
            "id": "o3",
            "orderNo": "LSX-2024-003",
            "productId": p3["id"] if p3 else "p3",
            "productCode": p3["code"] if p3 else "NOVO-VB-15-ABS",
            "productLine": p3["name"] if p3 else "Van bi NOVO 15 tay ABS",
            "customer": "Nội bộ",
            "targetQty": 800,
            "createdBy": "Nguyễn Văn An",
            "createdAt": "17/01/2024",
            "deadline": "2024-03-10",
            "priority": "normal",
            "status": "pending_approval",
            "pendingApproval": True,
            "note": "Chờ Quản đốc duyệt lệnh mới",
            "attachments": [],
            "boms": jobs_for("p3", 800, limit_parts=3) if p3 else [],
        },
        {
            "id": "o5",
            "orderNo": "LSX-2024-005",
            "productId": "p1",
            "productCode": p1["code"],
            "productLine": p1["name"],
            "customer": "Nội bộ",
            "targetQty": 500,
            "createdBy": "Nguyễn Văn An",
            "createdAt": "21/01/2024",
            "deadline": "2024-03-01",
            "priority": "high",
            "status": "in_progress",
            "pendingApproval": False,
            "note": "ĐMKT sheet Mẫu van: mỗi nguyên công = 1 job; tuần tự trong cùng linh kiện",
            "attachments": [],
            "boms": jobs_for("p1", 500, limit_parts=2, assign_first=True, progress=True),
        },
    ]
    if p4:
        orders.append(
            {
                "id": "o4",
                "orderNo": "LSX-2024-004",
                "productId": p4["id"],
                "productCode": p4["code"],
                "productLine": p4["name"],
                "customer": "Khách hàng C",
                "targetQty": 200,
                "createdBy": "Nguyễn Văn An",
                "createdAt": "20/01/2024",
                "deadline": "2024-02-20",
                "priority": "high",
                "status": "approved",
                "pendingApproval": False,
                "attachments": [],
                "boms": jobs_for("p4", 200, limit_parts=3),
            }
        )
    if p5:
        orders.append(
            {
                "id": "o6",
                "orderNo": "LSX-2024-006",
                "productId": p5["id"],
                "productCode": p5["code"],
                "productLine": p5["name"],
                "customer": "Nội bộ",
                "targetQty": 400,
                "createdBy": "Trần Thị Bình",
                "createdAt": "22/01/2024",
                "deadline": "2024-03-15",
                "priority": "normal",
                "status": "approved",
                "pendingApproval": False,
                "attachments": [],
                "boms": jobs_for("p5", 400, limit_parts=2),
            }
        )

    # Lệnh SX demo nằm ở shared/src/data/buildCatalogOrders.ts — không ghi lại file JSON 25-job.
    print(
        "products",
        len(products),
        "semis",
        len(semis),
        "processes",
        len(processes),
        "machines",
        len(machines),
        "orders skipped (buildCatalogOrders)",
    )


if __name__ == "__main__":
    main()
